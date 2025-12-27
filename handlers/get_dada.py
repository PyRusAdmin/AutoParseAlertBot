# -*- coding: utf-8 -*-
import os
from datetime import datetime

from aiogram import F
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, FSInputFile
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from database.database import User, create_keywords_model, create_groups_model
from locales.locales import get_text
from system.dispatcher import router


def create_excel_file(data: list, headers: list, filename: str, sheet_name: str) -> str:
    """
    Функция для создания Excel-файла с переданными данными.
    :param data: Список кортежей с данными (строки таблицы)
    :param headers: Список заголовков столбцов
    :param filename: Имя создаваемого файла
    :param sheet_name: Имя листа Excel
    :return: Путь к созданному файлу
    """
    # Создаём новый Excel-файл (рабочую книгу)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name  # Название листа

    # Настройки оформления заголовков таблицы
    header_font = Font(bold=True, color="FFFFFF", size=12)  # Жирный белый шрифт
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Синий фон
    header_alignment = Alignment(horizontal="center", vertical="center")  # Центрирование текста

    # Запись заголовков в первую строку
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Запись данных (начиная со второй строки)
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Автоматическая подгонка ширины столбцов под содержимое
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # ограничиваем ширину 50 символами
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Создаём папку exports, если её ещё нет
    exports_dir = "exports"
    os.makedirs(exports_dir, exist_ok=True)

    # Сохраняем файл
    filepath = os.path.join(exports_dir, filename)
    workbook.save(filepath)
    logger.info(f"Excel файл создан: {filepath}")

    return filepath


@router.message(F.text == "🔍 Список ключевых слов")
async def get_keywords_list(message: Message, state: FSMContext):
    """
    Экспорт списка ключевых слов пользователя в Excel-файл.
    Отправляет пользователю документ через Telegram.
    """
    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)

    logger.info(f"Пользователь {telegram_user.id} {telegram_user.username} запросил экспорт ключевых слов")

    # Получаем модель таблицы ключевых слов для данного пользователя
    KeywordsModel = create_keywords_model(user_id=telegram_user.id)

    # Проверяем, существует ли таблица
    if not KeywordsModel.table_exists():
        KeywordsModel.create_table()
        await message.answer(get_text(user.language, "no_keywords"))
        return

    # Извлекаем все ключевые слова
    keywords = list(KeywordsModel.select())

    if not keywords:
        await message.answer(get_text(user.language, "no_keywords"))
        return

    # Формируем список данных для записи в Excel
    data = []
    for idx, keyword in enumerate(keywords, start=1):
        data.append((idx, keyword.user_keyword))  # Номер и текст ключевого слова

    # Формируем имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"keywords_{telegram_user.id}_{timestamp}.xlsx"
    headers = ["№", "Ключевое слово / Keyword"]

    try:
        # Создаём Excel-файл
        filepath = create_excel_file(
            data=data,
            headers=headers,
            filename=filename,
            sheet_name="Keywords"
        )

        # Отправляем файл пользователю в Telegram
        document = FSInputFile(filepath)
        await message.answer_document(
            document=document,
            caption=f"📋 {get_text(user.language, 'keywords_export')}\n"
                    f"Всего записей: {len(data)}"
        )

        # Удаляем файл после отправки
        os.remove(filepath)
        logger.info(f"Файл ключевых слов отправлен и удалён: {filepath}")

    except Exception as e:
        logger.exception(f"Ошибка при создании Excel-файла с ключевыми словами: {e}")
        await message.answer(get_text(user.language, "export_error"))


@router.message(F.text == "🌐 Ссылки для отслеживания")
async def get_tracking_links_list(message: Message, state: FSMContext):
    """
    Экспорт списка ссылок (каналов/групп) для отслеживания в Excel-файл.
    Отправляет файл пользователю через Telegram.
    """
    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)

    logger.info(f"Пользователь {telegram_user.id} {telegram_user.username} запросил экспорт ссылок для отслеживания")

    # Получаем модель таблицы групп для данного пользователя
    GroupsModel = create_groups_model(user_id=telegram_user.id)

    # Проверяем, существует ли таблица
    if not GroupsModel.table_exists():
        GroupsModel.create_table()
        await message.answer(get_text(user.language, "no_tracking_links"))
        return

    # Извлекаем все записи (группы/каналы)
    groups = list(GroupsModel.select())

    if not groups:
        await message.answer(get_text(user.language, "no_tracking_links"))
        return

    # Формируем список данных для Excel
    data = []
    for idx, group in enumerate(groups, start=1):
        data.append((idx, group.username_chat_channel))  # Номер и username канала

    # Формируем имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tracking_links_{telegram_user.id}_{timestamp}.xlsx"
    headers = ["№", "Username канала/группы / Channel/Group Username"]

    try:
        # Создаём Excel-файл
        filepath = create_excel_file(
            data=data,
            headers=headers,
            filename=filename,
            sheet_name="Tracking Links"
        )

        # Отправляем файл пользователю
        document = FSInputFile(filepath)
        await message.answer_document(
            document=document,
            caption=f"🔗 {get_text(user.language, 'tracking_links_export')}\n"
                    f"Всего записей: {len(data)}"
        )

        # Удаляем файл после отправки
        os.remove(filepath)
        logger.info(f"Файл ссылок отправлен и удалён: {filepath}")

    except Exception as e:
        logger.exception(f"Ошибка при создании Excel-файла со ссылками: {e}")
        await message.answer(get_text(user.language, "export_error"))


def register_data_export_handlers():
    """
    Регистрация хэндлеров экспорта данных.
    (Можно вызвать при инициализации бота.)
    """
    router.message.register(get_keywords_list)
    router.message.register(get_tracking_links_list)
