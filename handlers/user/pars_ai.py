# -*- coding: utf-8 -*-
import io
import re
from datetime import datetime

from aiogram import F
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, ReplyKeyboardRemove, Message
from loguru import logger  # https://github.com/Delgan/loguru
from openpyxl import Workbook
from openpyxl.styles import Font

from ai.ai import get_groq_response, search_groups_in_telegram
from database.database import User, TelegramGroup
from keyboards.user.keyboards import back_keyboard, search_group_ai, get_categories_keyboard
from locales.locales import get_text
from states.states import MyStates, ExportStates
from system.dispatcher import router


def clean_group_name(name):
    """
    –û—á–∏—â–∞–µ—Ç –Ω–∞–∑–≤–∞–Ω–∏–µ –≥—Ä—É–ø–ø—ã –æ—Ç –Ω–∞—á–∞–ª—å–Ω—ã—Ö –Ω–æ–º–µ—Ä–æ–≤, —Å–∏–º–≤–æ–ª–æ–≤ –∏ –ª–∏—à–Ω–∏—Ö –ø—Ä–æ–±–µ–ª–æ–≤.

    –£–¥–∞–ª—è–µ—Ç —Å –Ω–∞—á–∞–ª–∞ —Å—Ç—Ä–æ–∫–∏ –ø–æ—Å–ª–µ–¥–æ–≤–∞—Ç–µ–ª—å–Ω–æ—Å—Ç–∏ –∏–∑ —Ü–∏—Ñ—Ä, —Ç–æ—á–µ–∫, —Ç–∏—Ä–µ, –∑–≤—ë–∑–¥–æ—á–µ–∫,
    —Å–∫–æ–±–æ–∫ –∏ –ø—Ä–æ–±–µ–ª–æ–≤, –∫–æ—Ç–æ—Ä—ã–µ —á–∞—Å—Ç–æ –ø—Ä–∏—Å—É—Ç—Å—Ç–≤—É—é—Ç –≤ –ø–µ—Ä–µ—á–∏—Å–ª–µ–Ω–Ω—ã—Ö —Å–ø–∏—Å–∫–∞—Ö.

    –ù–∞–ø—Ä–∏–º–µ—Ä, –ø—Ä–µ–æ–±—Ä–∞–∑—É–µ—Ç "1. –ì—Ä—É–ø–ø–∞ —Ä–∞–∑—Ä–∞–±–æ—Ç—á–∏–∫–æ–≤" –≤ "–ì—Ä—É–ø–ø–∞ —Ä–∞–∑—Ä–∞–±–æ—Ç—á–∏–∫–æ–≤".

    :param name : (str) –ò—Å—Ö–æ–¥–Ω–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –≥—Ä—É–ø–ø—ã.
    :return str: –û—á–∏—â–µ–Ω–Ω–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –≥—Ä—É–ø–ø—ã –±–µ–∑ –ø—Ä–µ—Ñ–∏–∫—Å–æ–≤.
    """
    cleaned = re.sub(r'^[\d\.\-\*\s\)\(\[\]]+', '', name).strip()
    return cleaned


# def generate_group_hash(username=None, name=None, link=None):
#     """
#     –ì–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç MD5-—Ö–µ—à –¥–ª—è —É–Ω–∏–∫–∞–ª—å–Ω–æ–π –∏–¥–µ–Ω—Ç–∏—Ñ–∏–∫–∞—Ü–∏–∏ –≥—Ä—É–ø–ø—ã –≤ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö.
#
#     –ò—Å–ø–æ–ª—å–∑—É–µ—Ç –æ–¥–∏–Ω –∏–∑ —Ç—Ä—ë—Ö –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤: username, link –∏–ª–∏ name (–≤ –ø–æ—Ä—è–¥–∫–µ –ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç–∞)
#     –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è —Ö–µ—à–∞, –∫–æ—Ç–æ—Ä—ã–π —Å–ª—É–∂–∏—Ç –ø–µ—Ä–≤–∏—á–Ω—ã–º –∫–ª—é—á–æ–º –≤ —Ç–∞–±–ª–∏—Ü–µ `TelegramGroup`.
#
#     –ü—Ä–∏–æ—Ä–∏—Ç–µ—Ç: username > link > name. –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –ø–µ—Ä–≤—ã–π –Ω–µ–ø—É—Å—Ç–æ–π –ø–∞—Ä–∞–º–µ—Ç—Ä.
#
#     :param username : (str, optional) –Æ–∑–µ—Ä–Ω–µ–π–º –≥—Ä—É–ø–ø—ã (–Ω–∞–ø—Ä–∏–º–µ—Ä, "@python_chat").
#     :param name : (str, optional) –ù–∞–∑–≤–∞–Ω–∏–µ –≥—Ä—É–ø–ø—ã.
#     :param link : (str, optional) –ü—Ä—è–º–∞—è —Å—Å—ã–ª–∫–∞ –Ω–∞ –≥—Ä—É–ø–ø—É (–Ω–∞–ø—Ä–∏–º–µ—Ä, "https://t.me/python_chat").
#     :return str: 32-—Å–∏–º–≤–æ–ª—å–Ω–∞—è hex-—Å—Ç—Ä–æ–∫–∞ MD5-—Ö–µ—à–∞.
#     """
#     if username:
#         return hashlib.md5(username.encode()).hexdigest()
#     elif link:
#         return hashlib.md5(link.encode()).hexdigest()
#     else:
#         return hashlib.md5(name.encode()).hexdigest()


# def determine_group_type(group_data):
#     """
#     –û–ø—Ä–µ–¥–µ–ª—è–µ—Ç —Ç–∏–ø Telegram-—á–∞—Ç–∞ –Ω–∞ –æ—Å–Ω–æ–≤–µ –µ–≥–æ –¥–∞–Ω–Ω—ã—Ö.
#
#     –ê–Ω–∞–ª–∏–∑–∏—Ä—É–µ—Ç —Å–ª–æ–≤–∞—Ä—å —Å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–µ–π –æ –≥—Ä—É–ø–ø–µ –∏ –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Å—Ç—Ä–æ–∫—É —Å —Ç–∏–ø–æ–º.
#
#     –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –≥—Ä—É–ø–ø—ã –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö.
#
#     - 'channel': –µ—Å–ª–∏ –µ—Å—Ç—å —Ñ–ª–∞–≥ is_channel.
#     - 'group': –µ—Å–ª–∏ –µ—Å—Ç—å username (–ø—Ä–µ–¥–ø–æ–ª–∞–≥–∞–µ—Ç, —á—Ç–æ —ç—Ç–æ –≥—Ä—É–ø–ø–∞).
#     - 'link': –≤–æ –≤—Å–µ—Ö –æ—Å—Ç–∞–ª—å–Ω—ã—Ö —Å–ª—É—á–∞—è—Ö.
#
#     :param group_data : (dict) –°–ª–æ–≤–∞—Ä—å —Å –¥–∞–Ω–Ω—ã–º–∏ –æ –≥—Ä—É–ø–ø–µ, –ø–æ–ª—É—á–µ–Ω–Ω—ã–π –∏–∑ –ø–æ–∏—Å–∫–∞.
#     :return str: –¢–∏–ø —á–∞—Ç–∞ ‚Äî 'channel', 'group' –∏–ª–∏ 'link'.
#     """
#     if 'is_channel' in group_data and group_data['is_channel']:
#         return 'channel'
#     elif 'username' in group_data and group_data['username']:
#         return 'group'
#     else:
#         return 'link'


def save_group_to_db(group_data):
    """
    –°–æ—Ö—Ä–∞–Ω—è–µ—Ç –∏–ª–∏ –æ–±–Ω–æ–≤–ª—è–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –≥—Ä—É–ø–ø–µ –≤ —Ü–µ–Ω—Ç—Ä–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–π –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö.

    –ò—Å–ø–æ–ª—å–∑—É–µ—Ç —Ö–µ—à –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ —Å—É—â–µ—Å—Ç–≤–æ–≤–∞–Ω–∏—è –≥—Ä—É–ø–ø—ã. –ü—Ä–∏ –Ω–∞–ª–∏—á–∏–∏ –æ–±–Ω–æ–≤–ª—è–µ—Ç –ø–æ–ª—è,
    –ø—Ä–∏ –æ—Ç—Å—É—Ç—Å—Ç–≤–∏–∏ ‚Äî —Å–æ–∑–¥–∞—ë—Ç –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å. –ü–æ–ª—è 'participants' –∏ 'description'
    –æ–±–Ω–æ–≤–ª—è—é—Ç—Å—è –ø—Ä–∏ –∫–∞–∂–¥–æ–º –Ω–∞—Ö–æ–∂–¥–µ–Ω–∏–∏ –≥—Ä—É–ø–ø—ã.

    –§—É–Ω–∫—Ü–∏—è —è–≤–ª—è–µ—Ç—Å—è —á–∞—Å—Ç—å—é –º–µ—Ö–∞–Ω–∏–∑–º–∞ deduplication –∏ –ø—Ä–µ–¥–æ—Ç–≤—Ä–∞—â–∞–µ—Ç –¥—É–±–ª–∏—Ä–æ–≤–∞–Ω–∏–µ –∑–∞–ø–∏—Å–µ–π.

    :param group_data : (dict) –°–ª–æ–≤–∞—Ä—å —Å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–µ–π –æ –≥—Ä—É–ø–ø–µ (–Ω–∞–∑–≤–∞–Ω–∏–µ, username, —É—á–∞—Å—Ç–Ω–∏–∫–∏ –∏ —Ç.–¥.).
    :return TelegramGroup or None: –≠–∫–∑–µ–º–ø–ª—è—Ä —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω–æ–π –º–æ–¥–µ–ª–∏ –∏–ª–∏ None –ø—Ä–∏ –æ—à–∏–±–∫–µ.
    :raise Exception: –õ–æ–≥–∏—Ä—É–µ—Ç—Å—è –ø—Ä–∏ –æ—à–∏–±–∫–∞—Ö —Ä–∞–±–æ—Ç—ã —Å –ë–î (–Ω–∞–ø—Ä–∏–º–µ—Ä, –Ω–∞—Ä—É—à–µ–Ω–∏–µ –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π).
    """
    try:
        telegram_id = group_data.get('telegram_id')
        group_hash = group_data.get('group_hash')
        name = group_data.get('name')
        username = group_data.get('username')
        description = group_data.get('description')
        participants = group_data.get('participants')
        category = group_data.get('category')
        group_type = group_data.get('group_type')
        language = group_data.get('language')
        link = group_data.get('link')
        date_added = datetime.now()
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ —É–∂–µ —Ç–∞–∫–∞—è –≥—Ä—É–ø–ø–∞
        existing = TelegramGroup.get_or_none(TelegramGroup.group_hash == group_hash)

        if existing:
            # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ
            existing.name = name
            existing.username = username
            existing.description = description
            existing.participants = participants
            existing.link = link
            existing.date_added = date_added
            existing.save()
            logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω–∞ –≥—Ä—É–ø–ø–∞: {group_data['name']}")
            return existing
        else:
            # –°–æ–∑–¥–∞—ë–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å
            new_group = TelegramGroup.create(
                telegram_id=telegram_id,
                group_hash=group_hash,
                name=name,
                username=username,
                description=description,
                participants=participants,
                category=category,
                group_type=group_type,
                language=language,
                link=link,
                date_added=date_added
            )
            logger.info(f"–î–æ–±–∞–≤–ª–µ–Ω–∞ –Ω–æ–≤–∞—è –≥—Ä—É–ø–ø–∞: {group_data['name']}")
            return new_group

    except Exception as e:
        logger.exception(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –≥—Ä—É–ø–ø—ã: {e}")
        return None


def format_summary_message(groups_count):
    """
    –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç HTML-—Å–æ–æ–±—â–µ–Ω–∏–µ —Å –∫—Ä–∞—Ç–∫–æ–π —Å–≤–æ–¥–∫–æ–π –æ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞—Ö –ø–æ–∏—Å–∫–∞.

    –í–∫–ª—é—á–∞–µ—Ç —Å—Ç–∞—Ç—É—Å –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è, –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –Ω–∞–π–¥–µ–Ω–Ω—ã—Ö –≥—Ä—É–ø–ø –∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –æ —Ñ–∞–π–ª–µ.

    –°–æ–æ–±—â–µ–Ω–∏–µ –æ—Ç–ø—Ä–∞–≤–ª—è–µ—Ç—Å—è –ø–µ—Ä–µ–¥ XLSX-—Ñ–∞–π–ª–æ–º.

    :param groups_count: (int) –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —É—Å–ø–µ—à–Ω–æ —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω—ã—Ö –∏ –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–Ω—ã—Ö –≥—Ä—É–ø–ø.
    :return: (str) –°–æ–æ–±—â–µ–Ω–∏–µ —Å HTML-—Ä–∞–∑–º–µ—Ç–∫–æ–π (—Ç–µ–≥–∏ <b>).
    """

    message = f"‚úÖ <b>–ü–æ–∏—Å–∫ –∑–∞–≤–µ—Ä—à—ë–Ω!</b>\n\n"
    message += f"üìä –ù–∞–π–¥–µ–Ω–æ –∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–æ: <b>{groups_count}</b> –≥—Ä—É–ø–ø/–∫–∞–Ω–∞–ª–æ–≤\n"
    message += f"üìÅ –†–µ–∑—É–ª—å—Ç–∞—Ç—ã –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω—ã –≤ Excel-—Ñ–∞–π–ª–µ"
    return message


def create_excel_file(groups):
    """
    –°–æ–∑–¥–∞—ë—Ç –±–∞–π—Ç–æ–≤—ã–π Excel-—Ñ–∞–π–ª (.xlsx) —Å –¥–∞–Ω–Ω—ã–º–∏ –æ –Ω–∞–π–¥–µ–Ω–Ω—ã—Ö –≥—Ä—É–ø–ø–∞—Ö –¥–ª—è –æ—Ç–ø—Ä–∞–≤–∫–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é.

    –°–æ–¥–µ—Ä–∂–∏—Ç –∫–æ–ª–æ–Ω–∫–∏: ID (Hash), –ù–∞–∑–≤–∞–Ω–∏–µ, Username, –û–ø–∏—Å–∞–Ω–∏–µ, –£—á–∞—Å—Ç–Ω–∏–∫–æ–≤,
    –ö–∞—Ç–µ–≥–æ—Ä–∏—è, –¢–∏–ø, –°—Å—ã–ª–∫–∞, –î–∞—Ç–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è.
    Username –ø—Ä–∏–≤–æ–¥–∏—Ç—Å—è –∫ —Ñ–æ—Ä–º–∞—Ç—É '@username'.

    :param groups: (list[TelegramGroup]) –°–ø–∏—Å–æ–∫ —ç–∫–∑–µ–º–ø–ª—è—Ä–æ–≤ –º–æ–¥–µ–ª–∏ TelegramGroup.
    :return: bytes ‚Äî —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ .xlsx —Ñ–∞–π–ª–∞ –≤ –ø–∞–º—è—Ç–∏.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "–†–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–æ–∏—Å–∫–∞"

    # –ó–∞–≥–æ–ª–æ–≤–∫–∏
    headers = [
        'ID (Hash)',
        '–ù–∞–∑–≤–∞–Ω–∏–µ',
        'Username',
        '–û–ø–∏—Å–∞–Ω–∏–µ',
        '–£—á–∞—Å—Ç–Ω–∏–∫–æ–≤',
        '–ö–∞—Ç–µ–≥–æ—Ä–∏—è',
        '–¢–∏–ø',
        '–°—Å—ã–ª–∫–∞',
        '–î–∞—Ç–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è'
    ]
    ws.append(headers)

    # –ñ–∏—Ä–Ω—ã–π —à—Ä–∏—Ñ—Ç –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # –î–∞–Ω–Ω—ã–µ
    for group in groups:
        username = group.username or ''
        if username:
            username = f"@{username.lstrip('@')}"

        ws.append([
            group.group_hash,
            group.name,
            username,
            group.description or '',
            group.participants,
            group.category or '',
            group.group_type,
            group.link,
            group.date_added.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # –ê–≤—Ç–æ–ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = min(length, 50)

    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.message(F.text == "üì• –í—Å—è –±–∞–∑–∞")
async def export_all_groups(message: Message, state: FSMContext):
    """–í—ã–¥–∞—ë—Ç CSV-—Ñ–∞–π–ª —Å–æ –≤—Å–µ–π –±–∞–∑–æ–π –¥–∞–Ω–Ω—ã—Ö –≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤."""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏—è
    try:
        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –∑–∞–ø–∏—Å–∏ –∏–∑ –±–∞–∑—ã
        groups = TelegramGroup.select()
        if not groups:
            await message.answer("üì≠ –ë–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö –ø—É—Å—Ç–∞.")
            return

        excel_bytes = create_excel_file(groups)
        document = BufferedInputFile(excel_bytes, filename="–í—Å—è_–±–∞–∑–∞.xlsx")
        await message.answer_document(
            document=document,
            caption=f"üì¶ –í—Å—è –±–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö Telegram-–≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤.\n\nüìä –í—Å–µ–≥–æ –∑–∞–ø–∏—Å–µ–π: {len(groups)}"
        )

    except Exception as e:
        await message.answer("‚ùå –ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ —Ñ–∞–π–ª–∞.")
        logger.exception(e)


@router.message(F.text == "üì• –ë–∞–∑–∞ –∫–∞–Ω–∞–ª–æ–≤")
async def export_channels(message: Message, state: FSMContext):
    """–í—ã–¥–∞—ë—Ç CSV-—Ñ–∞–π–ª —Å–æ –≤—Å–µ–π –±–∞–∑–æ–π –¥–∞–Ω–Ω—ã—Ö –≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤."""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏—è
    # –ü—É—Ç—å –∫ –≤—Ä–µ–º–µ–Ω–Ω–æ–º—É CSV-—Ñ–∞–π–ª—É
    try:
        # –ü–æ–ª—É—á–∞–µ–º —Ç–æ–ª—å–∫–æ –ö–ê–ù–ê–õ–´
        groups = TelegramGroup.select().where(
            TelegramGroup.group_type == '–ö–∞–Ω–∞–ª'
        )

        if not groups:
            await message.answer("üì≠ –ë–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö –ø—É—Å—Ç–∞.")
            return

        excel_bytes = create_excel_file(groups)
        document = BufferedInputFile(excel_bytes, filename="–ë–∞–∑–∞_–∫–∞–Ω–∞–ª–æ–≤.xlsx")
        await message.answer_document(
            document=document,
            caption=f"üì¶ –í—Å—è –±–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö Telegram-–≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤.\n\nüìä –í—Å–µ–≥–æ –∑–∞–ø–∏—Å–µ–π: {len(groups)}"
        )

    except Exception as e:
        await message.answer("‚ùå –ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ —Ñ–∞–π–ª–∞.")
        logger.exception(e)


@router.message(F.text == "üì• –ë–∞–∑–∞ –≥—Ä—É–ø–ø")
async def export_supergroups(message: Message, state: FSMContext):
    """–í—ã–¥–∞—ë—Ç CSV-—Ñ–∞–π–ª —Å–æ –≤—Å–µ–π –±–∞–∑–æ–π –¥–∞–Ω–Ω—ã—Ö –≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤."""
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏—è
    try:
        # –ü–æ–ª—É—á–∞–µ–º —Ç–æ–ª—å–∫–æ –°–£–ü–ï–†–ì–†–£–ü–ü–´
        groups = TelegramGroup.select().where(
            TelegramGroup.group_type == '–ì—Ä—É–ø–ø–∞ (—Å—É–ø–µ—Ä–≥—Ä—É–ø–ø–∞)'
        )
        if not groups:
            await message.answer("üì≠ –ë–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö –ø—É—Å—Ç–∞.")
            return

        excel_bytes = create_excel_file(groups)
        document = BufferedInputFile(excel_bytes, filename="–ë–∞–∑–∞_–≥—Ä—É–ø–ø.xlsx")
        await message.answer_document(
            document=document,
            caption=f"üì¶ –í—Å—è –±–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö Telegram-–≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤.\n\nüìä –í—Å–µ–≥–æ –∑–∞–ø–∏—Å–µ–π: {len(groups)}"
        )

    except Exception as e:
        await message.answer("‚ùå –ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ —Ñ–∞–π–ª–∞.")
        logger.exception(e)


@router.message(F.text == "üì• –ü–æ–ª—É—á–∏—Ç—å –±–∞–∑—É")
async def handle_enter_keyword_menu(message: Message, state: FSMContext):
    """
    –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –∑–∞–ø—Ä–æ—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –Ω–∞ –ø–æ–ª—É—á–µ–Ω–∏–µ –±–∞–∑—ã Telegram-–≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤.

    –û—Ç–æ–±—Ä–∞–∂–∞–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–æ–Ω–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å –æ–ø–∏—Å–∞–Ω–∏–µ–º –¥–æ—Å—Ç—É–ø–Ω—ã—Ö –¥–µ–π—Å—Ç–≤–∏–π:
    - üì• –ü–æ–ª—É—á–µ–Ω–∏–µ –≤—Å–µ–π –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö
    - üîô –í–æ–∑–≤—Ä–∞—Ç –≤ –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é

    –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –∫–∞–∫ –ø—Ä–æ–º–µ–∂—É—Ç–æ—á–Ω–æ–µ –º–µ–Ω—é –¥–ª—è –Ω–∞–≤–∏–≥–∞—Ü–∏–∏ –≤ —Ä–∞–∑–¥–µ–ª–µ –ø–æ–∏—Å–∫–∞.

    :param message: (Message) –í—Ö–æ–¥—è—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –æ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è.
    :param state: (FSMContext, optional) –ö–æ–Ω—Ç–µ–∫—Å—Ç —Å–æ—Å—Ç–æ—è–Ω–∏—è –∫–æ–Ω–µ—á–Ω–æ–≥–æ –∞–≤—Ç–æ–º–∞—Ç–∞ (–Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è, –Ω–æ –ø–µ—Ä–µ–¥–∞—ë—Ç—Å—è).
    :return: None
    """
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏—è
    text = (
        "üëã –î–æ–±—Ä–æ –ø–æ–∂–∞–ª–æ–≤–∞—Ç—å –≤ —Ä–µ–∂–∏–º –ø–æ–ª—É—á–µ–Ω–∏—è –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö!\n\n"
        "–í–æ—Ç —á—Ç–æ –≤—ã –º–æ–∂–µ—Ç–µ —Å–¥–µ–ª–∞—Ç—å:\n\n"

        "üîπ <b>üì• –ü–æ–ª—É—á–∏—Ç—å –≤—Å—é –±–∞–∑—É</b> ‚Äî –ø–æ–ª—É—á–∏—Ç–µ –ø–æ–ª–Ω—ã–π —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω—ã—Ö –≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤ –≤ —Ñ–æ—Ä–º–∞—Ç–µ Excel.\n"
        "üîπ <b>üì• –ü–æ–ª—É—á–∏—Ç—å –±–∞–∑—É –ö–∞–Ω–∞–ª–æ–≤</b> ‚Äî –ø–æ–ª—É—á–∏—Ç–µ —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω—ã—Ö –∫–∞–Ω–∞–ª–æ–≤ –≤ —Ñ–æ—Ä–º–∞—Ç–µ Excel.\n"
        "üîπ <b>üì• –ü–æ–ª—É—á–∏—Ç—å –±–∞–∑—É –ì—Ä—É–ø–ø (—Å—É–ø–µ—Ä–≥—Ä—É–ø–ø)</b> ‚Äî –ø–æ–ª—É—á–∏—Ç–µ —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω—ã—Ö —Å—É–ø–µ—Ä–≥—Ä—É–ø–ø –≤ —Ñ–æ—Ä–º–∞—Ç–µ Excel.\n"
        "üîπ <b>üì• –ü–æ–ª—É—á–∏—Ç—å –±–∞–∑—É –û–±—ã—á–Ω—ã—Ö —á–∞—Ç–æ–≤ (–≥—Ä—É–ø–ø—ã —Å—Ç–∞—Ä–æ–≥–æ —Ç–∏–ø–∞)</b> ‚Äî –ø–æ–ª—É—á–∏—Ç–µ —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö —Å–æ—Ö—Ä–∞–Ω—ë–Ω–Ω—ã—Ö –æ–±—ã—á–Ω—ã—Ö —á–∞—Ç–æ–≤ (–≥—Ä—É–ø–ø —Å—Ç–∞—Ä–æ–≥–æ —Ç–∏–ø–∞) –≤ —Ñ–æ—Ä–º–∞—Ç–µ Excel.\n"
        "üîπ –í—ã–±—Ä–∞—Ç—å –∫–∞—Ç–µ–≥–æ—Ä–∏—é –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –±–∞–∑—ã\n\n"

        "üî∏ –ù–∞–∂–º–∏—Ç–µ <b>üîô –ù–∞–∑–∞–¥</b>, —á—Ç–æ–±—ã –≤–µ—Ä–Ω—É—Ç—å—Å—è –≤ –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é."
    )
    await message.answer(
        text=text,
        reply_markup=search_group_ai(),
        parse_mode="HTML"
    )


@router.message(F.text == "–í—ã–±—Ä–∞—Ç—å –∫–∞—Ç–µ–≥–æ—Ä–∏—é")
async def start_category_export(message: Message, state: FSMContext):
    """
    –ó–∞–ø—É—Å–∫–∞–µ—Ç –ø—Ä–æ—Ü–µ—Å—Å –≤—ã–±–æ—Ä–∞ –∫–∞—Ç–µ–≥–æ—Ä–∏–∏ –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞.
    –ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç –∫–ª–∞–≤–∏–∞—Ç—É—Ä—É —Å –∫–∞—Ç–µ–≥–æ—Ä–∏—è–º–∏ –∏ –ø–µ—Ä–µ–≤–æ–¥–∏—Ç –≤ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –æ–∂–∏–¥–∞–Ω–∏—è –≤—ã–±–æ—Ä–∞.
    """
    await message.answer(
        "üìå –í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–µ–≥–æ—Ä–∏—é, –ø–æ –∫–æ—Ç–æ—Ä–æ–π —Ö–æ—Ç–∏—Ç–µ –ø–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –≥—Ä—É–ø–ø/–∫–∞–Ω–∞–ª–æ–≤:",
        reply_markup=get_categories_keyboard()
    )
    await state.set_state(ExportStates.waiting_for_category)


@router.message(ExportStates.waiting_for_category)
async def handle_category_selection(message: Message, state: FSMContext):
    """
    –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –≤—ã–±–æ—Ä –∫–∞—Ç–µ–≥–æ—Ä–∏–∏ –∏ —Ñ–æ—Ä–º–∏—Ä—É–µ—Ç —Ñ–∞–π–ª —Å–æ —Å–ø–∏—Å–∫–æ–º –≥—Ä—É–ø–ø.
    """
    selected_category = message.text.strip()

    # –ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –∫–Ω–æ–ø–∫—É "–ù–∞–∑–∞–¥"
    if selected_category == "üîô –ù–∞–∑–∞–¥":
        await message.answer("‚ùå –û—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=ReplyKeyboardRemove())
        await state.clear()
        return

    # –°–ø–∏—Å–æ–∫ –¥–æ–ø—É—Å—Ç–∏–º—ã—Ö –∫–∞—Ç–µ–≥–æ—Ä–∏–π (–¥–ª—è –∑–∞—â–∏—Ç—ã –æ—Ç —Ä—É—á–Ω–æ–≥–æ –≤–≤–æ–¥–∞)
    valid_categories = {
        "–ò–Ω–≤–µ—Å—Ç–∏—Ü–∏–∏",
        "–§–∏–Ω–∞–Ω—Å—ã –∏ –ª–∏—á–Ω—ã–π –±—é–¥–∂–µ—Ç",
        "–ö—Ä–∏–ø—Ç–æ–≤–∞–ª—é—Ç—ã –∏ –±–ª–æ–∫—á–µ–π–Ω",
        "–ë–∏–∑–Ω–µ—Å –∏ –ø—Ä–µ–¥–ø—Ä–∏–Ω–∏–º–∞—Ç–µ–ª—å—Å—Ç–≤–æ",
        "–ú–∞—Ä–∫–µ—Ç–∏–Ω–≥ –∏ –ø—Ä–æ–¥–≤–∏–∂–µ–Ω–∏–µ",
        "–¢–µ—Ö–Ω–æ–ª–æ–≥–∏–∏ –∏ IT",
        "–û–±—Ä–∞–∑–æ–≤–∞–Ω–∏–µ –∏ —Å–∞–º–æ—Ä–∞–∑–≤–∏—Ç–∏–µ",
        "–†–∞–±–æ—Ç–∞ –∏ –∫–∞—Ä—å–µ—Ä–∞",
        "–ù–µ–¥–≤–∏–∂–∏–º–æ—Å—Ç—å",
        "–ó–¥–æ—Ä–æ–≤—å–µ –∏ –º–µ–¥–∏—Ü–∏–Ω–∞",
        "–ü—É—Ç–µ—à–µ—Å—Ç–≤–∏—è",
        "–ê–≤—Ç–æ –∏ —Ç—Ä–∞–Ω—Å–ø–æ—Ä—Ç",
        "–®–æ–ø–ø–∏–Ω–≥ –∏ —Å–∫–∏–¥–∫–∏",
        "–†–∞–∑–≤–ª–µ—á–µ–Ω–∏—è –∏ –¥–æ—Å—É–≥",
        "–ü–æ–ª–∏—Ç–∏–∫–∞ –∏ –æ–±—â–µ—Å—Ç–≤–æ",
        "–ù–∞—É–∫–∞ –∏ –∏—Å—Å–ª–µ–¥–æ–≤–∞–Ω–∏—è",
        "–°–ø–æ—Ä—Ç –∏ —Ñ–∏—Ç–Ω–µ—Å",
        "–ö—É–ª–∏–Ω–∞—Ä–∏—è –∏ –µ–¥–∞",
        "–ú–æ–¥–∞ –∏ –∫—Ä–∞—Å–æ—Ç–∞",
        "–•–æ–±–±–∏ –∏ —Ç–≤–æ—Ä—á–µ—Å—Ç–≤–æ"
    }

    if selected_category not in valid_categories:
        await message.answer(
            "‚ö†Ô∏è –ù–µ–≤–µ—Ä–Ω–∞—è –∫–∞—Ç–µ–≥–æ—Ä–∏—è. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=get_categories_keyboard()
        )
        return

    # –ü–æ–ª—É—á–∞–µ–º –≥—Ä—É–ø–ø—ã –∏–∑ –±–∞–∑—ã
    groups = TelegramGroup.select().where(TelegramGroup.category == selected_category)
    group_count = groups.count()

    if group_count == 0:
        await message.answer(
            f"üì≠ –í –∫–∞—Ç–µ–≥–æ—Ä–∏–∏ ¬´{selected_category}¬ª –ø–æ–∫–∞ –Ω–µ—Ç –Ω–∏ –æ–¥–Ω–æ–π –≥—Ä—É–ø–ø—ã.",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.clear()
        return

    # === –°–æ–∑–¥–∞—ë–º Excel-—Ñ–∞–π–ª ===
    wb = Workbook()
    ws = wb.active
    ws.title = "–ì—Ä—É–ø–ø—ã"

    # –ó–∞–≥–æ–ª–æ–≤–∫–∏
    headers = ["Username", "–ù–∞–∑–≤–∞–Ω–∏–µ", "–û–ø–∏—Å–∞–Ω–∏–µ", "–¢–∏–ø", "–£—á–∞—Å—Ç–Ω–∏–∫–∏", "–°—Å—ã–ª–∫–∞"]
    ws.append(headers)

    # –ñ–∏—Ä–Ω—ã–π —à—Ä–∏—Ñ—Ç –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # –î–∞–Ω–Ω—ã–µ
    for group in groups:
        ws.append([
            group.username or "",
            group.name or "",
            group.description or "",
            group.group_type or "",
            group.participants or 0,
            group.link or ""
        ])

    # –ê–≤—Ç–æ–ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã –∫–æ–ª–æ–Ω–æ–∫ (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # –æ–≥—Ä–∞–Ω–∏—á–∏–º —à–∏—Ä–∏–Ω—É
        ws.column_dimensions[column].width = adjusted_width

    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –ø–∞–º—è—Ç—å
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Ñ–∞–π–ª
    file_name = f"groups_{selected_category.replace(' ', '_')}.xlsx"
    await message.answer_document(
        document=BufferedInputFile(
            file=output.getvalue(),
            filename=file_name
        ),
        caption=f"‚úÖ –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {group_count} –≥—Ä—É–ø–ø/–∫–∞–Ω–∞–ª–æ–≤ –ø–æ –∫–∞—Ç–µ–≥–æ—Ä–∏–∏:\n¬´{selected_category}¬ª",
        reply_markup=ReplyKeyboardRemove()
    )

    logger.info(f"–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {message.from_user.id} —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–ª Excel –ø–æ –∫–∞—Ç–µ–≥–æ—Ä–∏–∏: {selected_category}")
    await state.clear()


@router.message(F.text == "ü§ñ AI –ø–æ–∏—Å–∫")
async def ai_search(message: Message, state: FSMContext):
    """
    –û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã "üì• –ü–æ–ª—É—á–∏—Ç—å –±–∞–∑—É".

    –û—á–∏—â–∞–µ—Ç —Å–æ—Å—Ç–æ—è–Ω–∏–µ FSM, –ø–æ–ª—É—á–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è, –ª–æ–≥–∏—Ä—É–µ—Ç –¥–µ–π—Å—Ç–≤–∏–µ
    –∏ –∑–∞–ø—Ä–∞—à–∏–≤–∞–µ—Ç —É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –∫–ª—é—á–µ–≤–æ–µ —Å–ª–æ–≤–æ –¥–ª—è –ø–æ–∏—Å–∫–∞ –≥—Ä—É–ø–ø —á–µ—Ä–µ–∑ AI.
    –ü–µ—Ä–µ–≤–æ–¥–∏—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –æ–∂–∏–¥–∞–Ω–∏—è –≤–≤–æ–¥–∞ (MyStates.entering_keyword_ai_search).

    :param message: (Message) –í—Ö–æ–¥—è—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –æ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è.
    :param state: (FSMContext) –ö–æ–Ω—Ç–µ–∫—Å—Ç –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏–π, —Å–±—Ä–∞—Å—ã–≤–∞–µ—Ç—Å—è –ø—Ä–∏ –≤—Ö–æ–¥–µ.
    :return: None
    """
    await state.clear()  # –°–±—Ä–∞—Å—ã–≤–∞–µ—Ç —Å–æ—Å—Ç–æ—è–Ω–∏–µ

    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)

    logger.info(
        f"–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {telegram_user.id} {telegram_user.username} –ø–µ—Ä–µ—à–µ–ª –≤ –º–µ–Ω—é –ø–æ–∏—Å–∫–∞ –≥—Ä—É–ø–ø")

    await message.answer(
        get_text(user.language, "enter_keyword"),
        reply_markup=back_keyboard()
    )
    await state.set_state(MyStates.entering_keyword_ai_search)


@router.message(MyStates.entering_keyword_ai_search)
async def handle_enter_keyword(message: Message, state: FSMContext):
    """
    –û–±—Ä–∞–±–æ—Ç—á–∏–∫ –≤–≤–æ–¥–∞ –∫–ª—é—á–µ–≤–æ–≥–æ —Å–ª–æ–≤–∞ –¥–ª—è AI-–ø–æ–∏—Å–∫–∞ –≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤.

    –ü–æ–ª—É—á–∞–µ—Ç –∑–∞–ø—Ä–æ—Å –æ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è, –≥–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç –≤–∞—Ä–∏–∞–Ω—Ç—ã –Ω–∞–∑–≤–∞–Ω–∏–π —á–µ—Ä–µ–∑ Groq API,
    –∏—â–µ—Ç —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤—É—é—â–∏–µ –≥—Ä—É–ø–ø—ã –≤ Telegram, —Å–æ—Ö—Ä–∞–Ω—è–µ—Ç –∏—Ö –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö –∏ –æ—Ç–ø—Ä–∞–≤–ª—è–µ—Ç
    —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é –≤ –≤–∏–¥–µ XLSX-—Ñ–∞–π–ª–∞.

    –í –ø—Ä–æ—Ü–µ—Å—Å–µ –ø–æ–∫–∞–∑—ã–≤–∞–µ—Ç —Å—Ç–∞—Ç—É—Å "–ò—â—É...", —É–¥–∞–ª—è–µ—Ç –µ–≥–æ –ø–æ—Å–ª–µ –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è –∏ –æ—Ç–ø—Ä–∞–≤–ª—è–µ—Ç
    —Å–≤–æ–¥–∫—É –∏ —Ñ–∞–π–ª.

    –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –æ—à–∏–±–∫–∏ –∏ –ø—É—Å—Ç—ã–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã.

    - –ò—Å–ø–æ–ª—å–∑—É–µ—Ç `get_groq_response` –¥–ª—è –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –Ω–∞–∑–≤–∞–Ω–∏–π.
    - –ò—Å–ø–æ–ª—å–∑—É–µ—Ç `search_groups_in_telegram` –¥–ª—è –ø–æ–∏—Å–∫–∞ –≤ Telegram.
    - –†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Å–æ—Ö—Ä–∞–Ω—è—é—Ç—Å—è —á–µ—Ä–µ–∑ `save_group_to_db`.
    - –§–∞–π–ª —Å–æ–∑–¥–∞—ë—Ç—Å—è —á–µ—Ä–µ–∑ `create_excel_file` –∏ –æ—Ç–ø—Ä–∞–≤–ª—è–µ—Ç—Å—è –∫–∞–∫ –¥–æ–∫—É–º–µ–Ω—Ç.

    :param message: (Message) –í—Ö–æ–¥—è—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å –∫–ª—é—á–µ–≤—ã–º —Å–ª–æ–≤–æ–º.
    :param state: (FSMContext) –ö–æ–Ω—Ç–µ–∫—Å—Ç –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏–π, —Å–±—Ä–∞—Å—ã–≤–∞–µ—Ç—Å—è –ø–æ—Å–ª–µ –æ–±—Ä–∞–±–æ—Ç–∫–∏.
    :return: None

    Raises:
        Exception: –ü–µ—Ä–µ—Ö–≤–∞—Ç—ã–≤–∞–µ—Ç—Å—è –ª–æ–∫–∞–ª—å–Ω–æ, –ª–æ–≥–∏—Ä—É–µ—Ç—Å—è –∏ –ø—Ä–µ–æ–±—Ä–∞–∑—É–µ—Ç—Å—è –≤ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å—Å–∫–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ.
    """

    telegram_user = message.from_user
    user_input = message.text.strip()
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ –æ –Ω–∞—á–∞–ª–µ –ø–æ–∏—Å–∫–∞
    processing_msg = await message.answer("üîç –ò—â—É –≥—Ä—É–ø–ø—ã –∏ –∫–∞–Ω–∞–ª—ã...")

    try:
        # –ü–æ–ª—É—á–∞–µ–º –æ—Ç–≤–µ—Ç –æ—Ç AI
        answer = await get_groq_response(user_input)
        logger.info(f"–û—Ç–≤–µ—Ç –æ—Ç Groq: {answer}")

        # –†–∞–∑–±–∏–≤–∞–µ–º –æ—Ç–≤–µ—Ç –Ω–∞ —Å—Ç—Ä–æ–∫–∏ –∏ –æ—á–∏—â–∞–µ–º
        group_names = [clean_group_name(line) for line in answer.splitlines() if line.strip()]
        group_names = [name for name in group_names if len(name) > 2]
        logger.info(f"–ü–æ–ª—É—á–µ–Ω–æ {len(group_names)} –Ω–∞–∑–≤–∞–Ω–∏–π: {group_names}")

        saved_groups = []

        for group_name in group_names:
            # –ò—â–µ–º –≤ Telegram
            results = await search_groups_in_telegram([group_name])
            logger.info(f"–ù–∞–π–¥–µ–Ω–æ {len(results)} –≥—Ä—É–ø–ø –¥–ª—è '{group_name}'")

            # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –≤ –ë–î
            for group_data in results:
                saved_group = save_group_to_db(group_data)
                if saved_group:
                    saved_groups.append(saved_group)

        # –£–¥–∞–ª—è–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ –æ –ø–æ–∏—Å–∫–µ
        await processing_msg.delete()

        # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
        if saved_groups:

            # –°–æ–∑–¥–∞—ë–º Excel-—Ñ–∞–π–ª
            excel_bytes = create_excel_file(saved_groups)
            filename = f"telegram_groups_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_file = BufferedInputFile(excel_bytes, filename=filename)

            summary = format_summary_message(len(saved_groups))
            await message.answer(summary, parse_mode="HTML")
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º CSV —Ñ–∞–π–ª
            await message.answer_document(
                document=excel_file,
                caption=f"üìÑ –†–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–æ–∏—Å–∫–∞ –ø–æ –∑–∞–ø—Ä–æ—Å—É: <b>{user_input}</b>",
                parse_mode="HTML"
            )
            logger.info(f"–û—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ {len(saved_groups)} –≥—Ä—É–ø–ø –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {telegram_user.id} –≤ Excel —Ñ–∞–π–ª–µ")
        else:
            await message.answer(
                "‚ùå –ö —Å–æ–∂–∞–ª–µ–Ω–∏—é, –ø–æ –≤–∞—à–µ–º—É –∑–∞–ø—Ä–æ—Å—É –Ω–∏—á–µ–≥–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–∏–µ –∫–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞.",
                reply_markup=back_keyboard()
            )
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –∑–∞–ø—Ä–æ—Å–∞: {e}")
        await processing_msg.delete()
        await message.answer(
            "‚ùå –ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–∏—Å–∫–µ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â—ë —Ä–∞–∑.",
            reply_markup=back_keyboard()
        )
    await state.clear()  # –ó–∞–≤–µ—Ä—à–∞–µ–º —Ç–µ–∫—É—â–µ–µ —Å–æ—Å—Ç–æ—è–Ω–∏–µ –º–∞—à–∏–Ω—ã —Å–æ—Å—Ç–æ—è–Ω–∏—è


def register_handlers_pars_ai():
    """
    –†–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–µ—Ç –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ –¥–ª—è AI-–ø–æ–∏—Å–∫–∞ –∏ —ç–∫—Å–ø–æ—Ä—Ç–∞ Telegram-–≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤.

    –î–æ–±–∞–≤–ª—è–µ—Ç –≤ –º–∞—Ä—à—Ä—É—Ç–∏–∑–∞—Ç–æ—Ä (router) —Å–ª–µ–¥—É—é—â–∏–µ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏:
        1. search_menu ‚Äî –æ—Ç–æ–±—Ä–∞–∂–∞–µ—Ç –º–µ–Ω—é –ø–æ–∏—Å–∫–∞ –ø–æ –Ω–∞–∂–∞—Ç–∏—é –∫–Ω–æ–ø–∫–∏ "üì• –ü–æ–ª—É—á–∏—Ç—å –±–∞–∑—É".
        2. start_ai_search ‚Äî –∑–∞–ø—É—Å–∫–∞–µ—Ç –ø—Ä–æ—Ü–µ—Å—Å AI-–ø–æ–∏—Å–∫–∞ –ø–æ –Ω–∞–∂–∞—Ç–∏—é "ü§ñ AI –ø–æ–∏—Å–∫".
        3. process_ai_search_keyword ‚Äî –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –≤–≤–æ–¥ –∫–ª—é—á–µ–≤–æ–≥–æ —Å–ª–æ–≤–∞ –≤ —Å–æ—Å—Ç–æ—è–Ω–∏–∏ MyStates.entering_keyword_ai_search.
        4. export_all_groups ‚Äî —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç –≤—Å—é –±–∞–∑—É –≥—Ä—É–ø–ø –∏ –∫–∞–Ω–∞–ª–æ–≤ –≤ XLSX.
        5. export_channels ‚Äî —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç —Ç–æ–ª—å–∫–æ –∫–∞–Ω–∞–ª—ã.
        6. export_supergroups ‚Äî —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç —Ç–æ–ª—å–∫–æ —Å—É–ø–µ—Ä–≥—Ä—É–ø–ø—ã.
        7. export_legacy_groups ‚Äî —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç –æ–±—ã—á–Ω—ã–µ —á–∞—Ç—ã (–≥—Ä—É–ø–ø—ã —Å—Ç–∞—Ä–æ–≥–æ —Ç–∏–ø–∞).

    –≠—Ç–∏ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ –ø–æ–∑–≤–æ–ª—è—é—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é:
        - –ò—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å –ò–ò –¥–ª—è –ø–æ–∏—Å–∫–∞ —Ä–µ–ª–µ–≤–∞–Ω—Ç–Ω—ã—Ö Telegram-—áats –ø–æ –∫–ª—é—á–µ–≤–æ–º—É —Å–ª–æ–≤—É.
        - –ü–æ–ª—É—á–∞—Ç—å —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –≤ –≤–∏–¥–µ XLSX-—Ñ–∞–π–ª–∞.
        - –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å –≤—Å—é –∏–ª–∏ —á–∞—Å—Ç—å –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö –ø–æ —Ç–∏–ø–∞–º —á–∞—Ç–æ–≤.

    :return: None
    """
    router.message.register(handle_enter_keyword_menu, F.text == "üì• –ü–æ–ª—É—á–∏—Ç—å –±–∞–∑—É")
    router.message.register(ai_search, F.text == "ü§ñ AI –ø–æ–∏—Å–∫")
    router.message.register(export_all_groups, F.text == "üì• –í—Å—è –±–∞–∑–∞")
    router.message.register(export_channels, F.text == "üì• –ë–∞–∑–∞ –∫–∞–Ω–∞–ª–æ–≤")
    router.message.register(export_supergroups, F.text == "üì• –ë–∞–∑–∞ –≥—Ä—É–ø–ø")

    router.message.register(start_category_export, F.text == "–í—ã–±—Ä–∞—Ç—å –∫–∞—Ç–µ–≥–æ—Ä–∏—é")
    router.message.register(handle_category_selection, ExportStates.waiting_for_category)
